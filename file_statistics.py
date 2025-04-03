import os
import time
from datetime import datetime
from pathlib import Path
import humanize
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
import string

def get_file_stats(file_path):
    """Get statistics for a single file."""
    stats = os.stat(file_path)
    return {
        'size': stats.st_size,
        'created': datetime.fromtimestamp(stats.st_ctime),
        'modified': datetime.fromtimestamp(stats.st_mtime),
        'accessed': datetime.fromtimestamp(stats.st_atime),
        'extension': os.path.splitext(file_path)[1].lower() or 'no extension'
    }

def get_folder_size(folder_path):
    """Calculate total size of a folder and its contents."""
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            try:
                total_size += os.path.getsize(fp)
            except (OSError, FileNotFoundError):
                continue
    return total_size

def get_folder_depth(folder_path, base_path):
    """Calculate the depth of a folder relative to the base path."""
    rel_path = os.path.relpath(folder_path, base_path)
    return len(rel_path.split(os.sep)) if rel_path != '.' else 0

def format_path(path):
    """Format path to ensure proper drive letter format."""
    if path.endswith(':'):
        return path + '\\'
    return path

def scan_directory(directory):
    """Scan directory recursively and collect folder statistics."""
    folder_stats = []
    print(f"开始扫描目录: {directory}")
    print("这可能需要一些时间，请耐心等待...")
    
    for root, dirs, files in os.walk(directory):
        # Calculate folder depth
        depth = get_folder_depth(root, directory)
        
        # Get folder name
        folder_name = os.path.basename(root)
        if not folder_name:  # 处理根目录
            folder_name = root
        
        # Calculate folder size
        folder_size = get_folder_size(root)
        
        # Format path
        formatted_path = format_path(root)
        
        # Add to statistics
        folder_stats.append({
            'depth': depth,
            'name': folder_name,
            'path': formatted_path,
            'size': folder_size,
            'file_count': len(files)
        })
        
        # Print progress
        if len(folder_stats) % 100 == 0:
            print(f"已扫描 {len(folder_stats)} 个文件夹...")
    
    print(f"扫描完成！共发现 {len(folder_stats)} 个文件夹")
    
    # Sort by size in descending order
    folder_stats.sort(key=lambda x: x['size'], reverse=True)
    
    return folder_stats

def generate_excel_report(stats, output_file='folder_statistics.xlsx'):
    """Generate an Excel report from the folder statistics."""
    # 确保文件名以.xlsx结尾
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "文件夹统计"
        
        # 设置表头
        headers = ['层级', '文件夹名称', '完整路径', '大小', '文件数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据
        for row, stat in enumerate(stats, 2):
            ws.cell(row=row, column=1, value=stat['depth'])
            ws.cell(row=row, column=2, value=stat['name'])
            
            # Add hyperlink for folder path
            path_cell = ws.cell(row=row, column=3, value=stat['path'])
            path_cell.hyperlink = stat['path']
            path_cell.font = Font(color="0000FF", underline="single")
            
            ws.cell(row=row, column=4, value=humanize.naturalsize(stat['size']))
            ws.cell(row=row, column=5, value=stat['file_count'])
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # 尝试保存文件
        try:
            wb.save(output_file)
            print(f"报告已成功生成: {output_file}")
        except PermissionError:
            # 如果文件被占用，尝试使用新的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_output_file = f'folder_statistics_{timestamp}.xlsx'
            print(f"原文件可能被占用，尝试保存为新文件: {new_output_file}")
            wb.save(new_output_file)
            print(f"报告已成功生成: {new_output_file}")
        finally:
            # 确保工作簿被关闭
            wb.close()
            
    except Exception as e:
        print(f"生成Excel报告时出错: {str(e)}")
        print("请确保:")
        print("1. Excel文件未被其他程序打开")
        print("2. 您有权限写入该目录")
        print("3. 磁盘空间充足")
        return False
    
    return True

def get_available_drives():
    """Get all available drives on the system."""
    drives = []
    for letter in string.ascii_uppercase:
        drive = f"{letter}:\\"
        if os.path.exists(drive):
            drives.append(drive)
    return drives

def main():
    """Main function to run the folder statistics program."""
    # Get all available drives
    drives = get_available_drives()
    if not drives:
        print("错误：未找到任何可用的驱动器！")
        return
    
    print(f"找到以下驱动器: {', '.join(drives)}")
    all_stats = []
    
    try:
        # Scan each drive
        for drive in drives:
            print(f"\n开始扫描驱动器: {drive}")
            try:
                # Scan directory and collect statistics
                stats = scan_directory(drive)
                all_stats.extend(stats)
            except Exception as e:
                print(f"扫描驱动器 {drive} 时出错: {str(e)}")
                continue
        
        if not all_stats:
            print("错误：未能收集到任何文件夹统计信息！")
            return
        
        # Generate Excel report
        output_file = 'folder_statistics.xlsx'
        if not generate_excel_report(all_stats, output_file):
            print("生成报告失败，请检查以上错误信息。")
            return
        
    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        print("请检查是否有足够的权限和磁盘空间。")
        return

if __name__ == "__main__":
    main() 